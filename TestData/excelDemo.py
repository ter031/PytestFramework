import openpyxl
book = openpyxl.load_workbook("C:\\Users\\deepa\\OneDrive\\Desktop\\PythonDemo.xlsx")

sheet = book.active
Dict ={}
#print(sheet.cell(row=2, column=2).value)
sheet.cell(row=2, column=2).value = "Pawan"
#print(sheet.cell(row=2, column=2).value)
#print(sheet.max_row)
#print(sheet.max_column)
#print(sheet["B3"].value)

for i in range(2,sheet.max_row+1):  #to get rows
    #if sheet.cell(row=i,column=1).value == "testcase4":
    for j in range(2,sheet.max_column+1):  #to get columns
        Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
    print(Dict)

